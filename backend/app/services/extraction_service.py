import io
import re
from typing import List, Dict, Any, Tuple


class ExtractionService:
    @staticmethod
    def extract_text_from_pdf(content: bytes) -> List[Dict[str, Any]]:
        """
        Extracts structured page blocks with text, page numbers, and bounding box coordinates.
        Uses PyMuPDF (fitz) if installed, otherwise pure text fallback parser.
        """
        pages = []
        try:
            import fitz  # PyMuPDF
            doc = fitz.open(stream=content, filetype="pdf")
            for page_num, page in enumerate(doc, start=1):
                page_text = page.get_text("text")
                blocks = page.get_text("blocks")
                block_list = []
                for b in blocks:
                    # b format: (x0, y0, x1, y1, "text", block_no, block_type)
                    if len(b) >= 5:
                        block_list.append({
                            "text": b[4].strip(),
                            "bbox": {"x0": float(b[0]), "y0": float(b[1]), "x1": float(b[2]), "y1": float(b[3])}
                        })
                pages.append({
                    "page_number": page_num,
                    "text": page_text,
                    "blocks": block_list
                })
        except Exception:
            # Fallback pure text extraction
            text_str = content.decode("utf-8", errors="ignore")
            pages.append({
                "page_number": 1,
                "text": text_str,
                "blocks": [{"text": text_str, "bbox": {"x0": 0.0, "y0": 0.0, "x1": 100.0, "y1": 100.0}}]
            })
        return pages

    @staticmethod
    def extract_text_from_docx(content: bytes) -> List[Dict[str, Any]]:
        """Extracts text sections from DOCX files."""
        text_lines = []
        try:
            import docx
            doc = docx.Document(io.BytesIO(content))
            for p in doc.paragraphs:
                if p.text.strip():
                    text_lines.append(p.text.strip())
        except Exception:
            text_lines = [content.decode("utf-8", errors="ignore")]

        full_text = "\n".join(text_lines)
        return [{
            "page_number": 1,
            "text": full_text,
            "blocks": [{"text": full_text, "bbox": {"x0": 0.0, "y0": 0.0, "x1": 100.0, "y1": 100.0}}]
        }]

    @staticmethod
    def extract_text_from_xlsx(content: bytes) -> List[Dict[str, Any]]:
        """Extracts tabular data from XLSX files."""
        text_lines = []
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    row_vals = [str(cell) for cell in row if cell is not None]
                    if row_vals:
                        text_lines.append(" | ".join(row_vals))
        except Exception:
            text_lines = [content.decode("utf-8", errors="ignore")]

        full_text = "\n".join(text_lines)
        return [{
            "page_number": 1,
            "text": full_text,
            "blocks": [{"text": full_text, "bbox": {"x0": 0.0, "y0": 0.0, "x1": 100.0, "y1": 100.0}}]
        }]

    def extract_document(self, filename: str, content: bytes, content_type: str) -> Tuple[List[Dict[str, Any]], str]:
        ext = filename.lower()
        if ext.endswith(".pdf") or content_type == "application/pdf":
            return self.extract_text_from_pdf(content), "TEXT_PARSER"
        elif ext.endswith(".docx"):
            return self.extract_text_from_docx(content), "TEXT_PARSER"
        elif ext.endswith(".xlsx"):
            return self.extract_text_from_xlsx(content), "TEXT_PARSER"
        else:
            text_str = content.decode("utf-8", errors="ignore")
            return [{
                "page_number": 1,
                "text": text_str,
                "blocks": [{"text": text_str, "bbox": {"x0": 0.0, "y0": 0.0, "x1": 100.0, "y1": 100.0}}]
            }], "TEXT_PARSER"


extraction_service = ExtractionService()
